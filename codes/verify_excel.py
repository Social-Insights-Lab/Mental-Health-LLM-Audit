#!/usr/bin/env python3
"""
Quick verification of created Excel files without version conflicts
"""

import json
import openpyxl
from collections import Counter

def verify_excel_files():
    """Verify the structure of created Excel files"""
    
    annotation_file = "/home/ad2688/Research/mental_disorder/results_new/human_annotation/human_annotation_sample_refined.xlsx"
    comparison_file = "/home/ad2688/Research/mental_disorder/results_new/human_annotation/llm_labels_comparison_refined.xlsx"
    
    print("VERIFYING ANNOTATION FILES")
    print("=" * 50)
    
    # Check annotation file
    wb1 = openpyxl.load_workbook(annotation_file)
    ws1 = wb1.active
    
    print(f"\n📄 Annotation File:")
    print(f"  Rows: {ws1.max_row}")
    print(f"  Columns: {ws1.max_column}")
    
    # Get header row
    headers1 = [cell.value for cell in ws1[1]]
    print(f"  Headers: {headers1}")
    
    # Count entries per filename
    filename_counts = Counter()
    for row in range(2, ws1.max_row + 1):
        filename = ws1[f'A{row}'].value
        if filename:
            filename_counts[filename] += 1
    
    print(f"  Entries per LLM:")
    for llm, count in sorted(filename_counts.items()):
        print(f"    {llm}: {count}")
    
    # Check comparison file
    wb2 = openpyxl.load_workbook(comparison_file)
    ws2 = wb2.active
    
    print(f"\n📄 Comparison File:")
    print(f"  Rows: {ws2.max_row}")
    print(f"  Columns: {ws2.max_column}")
    
    # Get header row
    headers2 = [cell.value for cell in ws2[1]]
    print(f"  Headers: {headers2}")
    
    # Count majority labels
    majority_counts = Counter()
    for row in range(2, ws2.max_row + 1):
        majority_label = ws2[f'H{row}'].value  # Assuming Majority_Label is column H
        if majority_label:
            majority_counts[majority_label] += 1
    
    print(f"  Majority label distribution:")
    for label, count in sorted(majority_counts.items()):
        print(f"    {label}: {count}")
    
    print(f"\n✅ Verification completed!")
    print(f"📊 Total entries: {ws1.max_row - 1}")
    print(f"🎯 Target per LLM: 100")
    print(f"🎯 Expected total: 1200")

if __name__ == "__main__":
    verify_excel_files()